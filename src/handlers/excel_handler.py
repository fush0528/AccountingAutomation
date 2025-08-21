from typing import List, Optional
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..models import AccountingEntry


class ExcelHandler:
    """負責處理 Excel 檔案的讀寫操作"""

    def __init__(self, file_path: str):
        """
        初始化 Excel 處理器
        Args:
            file_path: Excel 檔案路徑
        """
        self.file_path = file_path
        self.workbook: Optional[Workbook] = None
        self.worksheet: Optional[Worksheet] = None
        self.headers = [
            '年份', '月份', '日期', '時間',
            '平台', '商品名稱', '訂單數量',
            '銷售總額', '平台費用', '實收金額',
            '需要發票', '應稅'
        ]

    def load_workbook(self) -> bool:
        """載入 Excel 檔案"""
        try:
            self.workbook = load_workbook(self.file_path)
            self.worksheet = self.workbook.active
            return self._validate_workbook()
        except FileNotFoundError:
            print(f"找不到檔案：{self.file_path}")
            return False
        except Exception as e:
            print(f"載入工作簿時發生錯誤: {e}")
            return False

    def save_workbook(self) -> bool:
        """儲存 Excel 檔案"""
        try:
            if self.workbook:
                self.workbook.save(self.file_path)
                return True
            return False
        except Exception as e:
            print(f"儲存工作簿時發生錯誤: {e}")
            return False

    def read_entries(self) -> List[AccountingEntry]:
        """讀取所有記帳項目"""
        entries = []
        if not self.worksheet:
            return entries

        # 跳過標題列
        for row in self.worksheet.iter_rows(min_row=2, values_only=True):
            if not any(row):  # 跳過空行
                continue
            
            try:
                entry_dict = {
                    'year': str(row[0]),
                    'month': str(row[1]),
                    'day': str(row[2]),
                    'time': str(row[3]),
                    'platform': str(row[4]),
                    'product_name': str(row[5]),
                    'order_quantity': int(row[6]),
                    'total_sales': float(row[7]),
                    'platform_fee': float(row[8]),
                    'actual_income': float(row[9]),
                    'invoice_required': bool(row[10]),
                    'taxable': bool(row[11])
                }
                entry = AccountingEntry.from_dict(entry_dict)
                if entry.validate():
                    entries.append(entry)
            except Exception as e:
                print(f"讀取記帳項目時發生錯誤: {e}")
                continue

        return entries

    def add_entry(self, entry: AccountingEntry) -> bool:
        """新增記帳項目"""
        if not self.worksheet or not entry.validate():
            return False

        try:
            row_values = [
                entry.year, entry.month, entry.day, entry.time,
                entry.platform, entry.product_name,
                entry.order_quantity, entry.total_sales, entry.platform_fee,
                entry.actual_income, entry.invoice_required, entry.taxable
            ]
            self.worksheet.append(row_values)
            return True
        except Exception as e:
            print(f"新增記帳項目時發生錯誤: {e}")
            return False

    def update_entry(self, row_index: int, entry: AccountingEntry) -> bool:
        """更新指定的記帳項目"""
        if not self.worksheet or not entry.validate():
            return False

        try:
            # Excel 的列索引從 1 開始
            row_values = [
                entry.year, entry.month, entry.day, entry.time,
                entry.platform, entry.product_name,
                entry.order_quantity, entry.total_sales, entry.platform_fee,
                entry.actual_income, entry.invoice_required, entry.taxable
            ]
            for col, value in enumerate(row_values, start=1):
                self.worksheet.cell(row=row_index, column=col, value=value)
            return True
        except Exception as e:
            print(f"更新記帳項目時發生錯誤: {e}")
            return False

    def delete_entry(self, row_index: int) -> bool:
        """刪除指定的記帳項目"""
        if not self.worksheet:
            return False

        try:
            self.worksheet.delete_rows(row_index)
            return True
        except Exception as e:
            print(f"刪除記帳項目時發生錯誤: {e}")
            return False

    def get_entry_by_index(self, row_index: int) -> Optional[AccountingEntry]:
        """取得指定索引的記帳項目"""
        if not self.worksheet:
            return None

        try:
            row = tuple(self.worksheet.iter_rows(
                min_row=row_index,
                max_row=row_index,
                values_only=True
            ))[0]

            entry_dict = {
                'year': str(row[0]),
                'month': str(row[1]),
                'day': str(row[2]),
                'time': str(row[3]),
                'platform': str(row[4]),
                'product_name': str(row[5]),
                'order_quantity': int(row[6]),
                'total_sales': float(row[7]),
                'platform_fee': float(row[8]),
                'actual_income': float(row[9]),
                'invoice_required': bool(row[10]),
                'taxable': bool(row[11])
            }
            entry = AccountingEntry.from_dict(entry_dict)
            return entry if entry.validate() else None
        except Exception as e:
            print(f"取得記帳項目時發生錯誤: {e}")
            return None

    def _validate_workbook(self) -> bool:
        """驗證工作簿格式是否正確"""
        if not self.worksheet:
            return False

        try:
            # 檢查是否為空白工作表
            if self.worksheet.max_row < 1:
                # 如果是空白的，加入標題列
                self.worksheet.append(self.headers)
                return True

            # 檢查標題列
            headers = [cell.value for cell in next(self.worksheet.iter_rows(max_row=1))]
            if not all(header == expected for header, expected in zip(headers, self.headers)):
                # 如果標題不符合預期，清空工作表並加入正確的標題
                self.worksheet.delete_rows(1, self.worksheet.max_row)
                self.worksheet.append(self.headers)
            return True
        except Exception as e:
            print(f"驗證工作簿時發生錯誤: {e}")
            return False