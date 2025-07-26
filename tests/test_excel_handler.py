import unittest
from datetime import datetime
import os
from openpyxl import Workbook

from src.models import AccountingEntry
from src.handlers import ExcelHandler

class TestExcelHandler(unittest.TestCase):
    """ExcelHandler 類別的單元測試"""

    def setUp(self):
        """設定測試環境"""
        self.test_file = "test_accounting.xlsx"
        self.test_date = datetime.now()
        
        # 建立測試用的 Excel 檔案
        wb = Workbook()
        ws = wb.active
        headers = [
            'date', 'platform', 'product_name', 'order_quantity',
            'total_sales', 'platform_fee', 'actual_income',
            'invoice_required', 'taxable'
        ]
        ws.append(headers)
        wb.save(self.test_file)
        
        self.handler = ExcelHandler(self.test_file)
        self.test_entry = AccountingEntry(
            date=self.test_date,
            platform="蝦皮",
            product_name="測試商品",
            order_quantity=1,
            total_sales=100.0,
            platform_fee=10.0
        )

    def tearDown(self):
        """清理測試環境"""
        if os.path.exists(self.test_file):
            os.remove(self.test_file)

    def test_load_workbook(self):
        """測試載入工作簿"""
        self.assertTrue(self.handler.load_workbook())
        self.assertIsNotNone(self.handler.workbook)
        self.assertIsNotNone(self.handler.worksheet)

    def test_add_and_read_entry(self):
        """測試新增和讀取記帳項目"""
        self.handler.load_workbook()
        
        # 測試新增項目
        self.assertTrue(self.handler.add_entry(self.test_entry))
        self.assertTrue(self.handler.save_workbook())
        
        # 測試讀取項目
        entries = self.handler.read_entries()
        self.assertEqual(len(entries), 1)
        entry = entries[0]
        
        self.assertEqual(entry.date.date(), self.test_date.date())
        self.assertEqual(entry.platform, "蝦皮")
        self.assertEqual(entry.product_name, "測試商品")
        self.assertEqual(entry.order_quantity, 1)
        self.assertEqual(entry.total_sales, 100.0)
        self.assertEqual(entry.platform_fee, 10.0)
        self.assertEqual(entry.actual_income, 90.0)

    def test_update_entry(self):
        """測試更新記帳項目"""
        self.handler.load_workbook()
        self.handler.add_entry(self.test_entry)
        
        # 建立更新後的項目
        updated_entry = AccountingEntry(
            date=self.test_date,
            platform="蝦皮",
            product_name="更新商品",
            order_quantity=2,
            total_sales=200.0,
            platform_fee=20.0
        )
        
        # 測試更新第一個項目（列索引為2，因為第1列是標題）
        self.assertTrue(self.handler.update_entry(2, updated_entry))
        self.assertTrue(self.handler.save_workbook())
        
        # 驗證更新結果
        entries = self.handler.read_entries()
        self.assertEqual(len(entries), 1)
        entry = entries[0]
        self.assertEqual(entry.product_name, "更新商品")
        self.assertEqual(entry.order_quantity, 2)
        self.assertEqual(entry.total_sales, 200.0)

    def test_delete_entry(self):
        """測試刪除記帳項目"""
        self.handler.load_workbook()
        self.handler.add_entry(self.test_entry)
        
        # 確認項目已新增
        entries = self.handler.read_entries()
        self.assertEqual(len(entries), 1)
        
        # 測試刪除第一個項目（列索引為2，因為第1列是標題）
        self.assertTrue(self.handler.delete_entry(2))
        self.assertTrue(self.handler.save_workbook())
        
        # 驗證刪除結果
        entries = self.handler.read_entries()
        self.assertEqual(len(entries), 0)

    def test_get_entry_by_index(self):
        """測試取得指定索引的記帳項目"""
        self.handler.load_workbook()
        self.handler.add_entry(self.test_entry)
        
        # 測試取得第一個項目（列索引為2，因為第1列是標題）
        entry = self.handler.get_entry_by_index(2)
        self.assertIsNotNone(entry)
        self.assertEqual(entry.platform, "蝦皮")
        self.assertEqual(entry.product_name, "測試商品")
        
        # 測試取得不存在的索引
        entry = self.handler.get_entry_by_index(999)
        self.assertIsNone(entry)


if __name__ == '__main__':
    unittest.main()